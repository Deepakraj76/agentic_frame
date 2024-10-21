from app.Model_Layer.models.autogen_ai import autogen_main
from app.Application_Layer.utils.common_utils import doc_text
import pandas as pd
import mysql.connector
from jinja2 import Template
from app.Application_Layer.services import db_state
from app.Data_Layer.repository.paths import getPaths
import os
import pandas as pd
from openpyxl import load_workbook

def query_concat(columns, selected_tables):
    print("SELECTED TABLE :",selected_tables)
    genrate_query = ""
    for i, col in enumerate(columns):

        if i == len(columns) - 1:  # Last element
            genrate_query += f"ROUTINE_DEFINITION LIKE '%{col}%'"
        else:
            genrate_query += f"ROUTINE_DEFINITION LIKE '%{col}%' OR "


    template_for_choice = """
    {% if db_type == 'mysql' %}
    -- Create a temporary table to store the result of dependent objects
    CREATE TEMPORARY TABLE temp_sp_list (
        dep_type VARCHAR(255)
    );
    {% elif db_type == 'mssql' %}
    SELECT 
            sys.objects.name AS ObjectName, 
            sys.objects.type AS ObjectType, 
            sys.sql_modules.definition AS Definition
        FROM sys.objects
        INNER JOIN sys.sql_modules 
        ON sys.objects.object_id = sys.sql_modules.object_id
        WHERE {genrate_query}
        AND sys.sql_modules.definition LIKE '{{selected_tables}}'
        ORDER BY sys.objects.name
    {% endif %}
    """

    query1_template = Template(template_for_choice)
    create_temp_table = query1_template.render(db_type=db_state.db_type, genrate_query=genrate_query, selected_tables=selected_tables)

    insert_sp_list = f"""
    INSERT INTO temp_sp_list (dep_type)
        SELECT ROUTINE_NAME FROM information_schema.ROUTINES
        WHERE ROUTINE_SCHEMA = '{db_state.db_name}' AND ROUTINE_DEFINITION LIKE '%{selected_tables}%';;
    """
    print(create_temp_table,"LLLL",insert_sp_list)
    return create_temp_table,insert_sp_list




def chunk_sp_list(definition_list, column_list, impact_table_json, selected_tables):
    object_list = []

    base_path, uploads_path, outputs_path = getPaths()

    # File path for the Excel file
    excel_file_path = f"{outputs_path}/impact_files/impact_sp_and_views.xlsx"

    sp_temp_excel = []

    for i in range(len(definition_list)):
        object_name = definition_list[i][0]
        object_type = definition_list[i][1]
        sp_definition = definition_list[i][2]
        object_list.append({"ObjectName": object_name, "ObjectType": object_type})

        prompt_input = ""
        for column in column_list:
            impact_list = [j for j in impact_table_json if j['Impact_column'] == column]
            prompt_input += f"""
                Input Format:
                ObjectName: {object_name}
                TableName: {selected_tables}
                Impact Column: {column}
                Impact Changes: {impact_list[0]['Description']}
                """

        prompt_input += f"""
        Objects: 
            {sp_definition}
        """
        print(prompt_input)

        impact_analysis_output = autogen_main(assistant_name="Database Administrator (DBA)",
                                              assistant_sys_msg="You are a helpful AI assistant you can find what I expect",
                                              agent_name="Database Administrator (DBA) Agent",
                                              agent_sys_message="You are expertise in Database. So you could assist me.",
                                              prompt=prompt_input)

        sp_temp_excel.append(impact_analysis_output)

    # Function to append dataframe to Excel file
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None):
        # Create a workbook and sheet if the file doesn't exist
        if not os.path.exists(filename):
            df.to_excel(filename, sheet_name=sheet_name, index=False)
            return

        # Load existing workbook
        book = load_workbook(filename)
        
        # Create a new sheet if it doesn't exist
        if sheet_name not in book.sheetnames:
            book.create_sheet(sheet_name)
        
        # Select the sheet
        sheet = book[sheet_name]
        
        # Find the next empty row
        if startrow is None:
            startrow = sheet.max_row

        # Write the dataframe to the sheet
        for r, row in enumerate(df.values, start=startrow + 1):
            for c, value in enumerate(row, start=1):
                sheet.cell(row=r, column=c, value=value)

        # Write the column headers if the sheet is empty
        if startrow == 0:
            for c, value in enumerate(df.columns.values, start=1):
                sheet.cell(row=1, column=c, value=value)

        # Save the workbook
        book.save(filename)

    # Append object_list to Excel
    append_df_to_excel(excel_file_path, pd.DataFrame(object_list), sheet_name="ObjectsList")

    # Call doc_text function to create the docx file
    doc_text(sp_temp_excel, selected_tables)

    return sp_temp_excel, object_list


def find_sp_definition(selected_tables, impact_table_json):
    columns = [i['Impact_column'] for i in impact_table_json]
    print(columns,selected_tables)
    for selected_tables in selected_tables:
        query1,query2 = query_concat(columns, selected_tables)

        connection = db_state.db_con
        cursor = db_state.db_cursor

        if connection.is_connected():
            print("Connected to MySQL database")

            cursor = connection.cursor()
        # Execute the first query
        cursor.execute(query1)
        connection.commit()
        temp_result = cursor.fetchall()  # Fetch all results from the first query
        # Check if the temp_result is empty or not
        if(db_state.db_type == 'mysql'):
            cursor.execute(query2)
            connection.commit()
        # Fetch all results from the second query
        temp_result = cursor.fetchall()
        print(temp_result)

        select_query = f"""
        SELECT ROUTINE_NAME AS ObjectName, ROUTINE_TYPE AS ObjectType, ROUTINE_DEFINITION AS Definition
        FROM information_schema.ROUTINES
        WHERE ROUTINE_SCHEMA = '{db_state.db_name}'
        AND ROUTINE_NAME IN (SELECT dep_type FROM temp_sp_list);
        """
        cursor.execute(select_query)
        definition_list = cursor.fetchall()

        # Drop the temporary table (only commit after this since it's an action query)
        cursor.execute("DROP TEMPORARY TABLE IF EXISTS temp_sp_list;")
        connection.commit()  # Commit only after DML queries like DROP

        cursor.close()  # Close the cursor after use

        sp_temp_result, object_list = chunk_sp_list(definition_list, columns, impact_table_json, selected_tables)
    return "final_sp"
